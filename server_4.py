from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
from pathlib import Path
import os, threading, time

app = Flask(__name__)
CORS(app)

BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "ZAJEBISTE_DANE.xlsx"
DEFAULT_SHEET = "Sheet1"

# 👇 prawdziwe nagłówki z Twojego pliku
HEADERS = [
    "SELECTOR","FOLDER","ADDED","LABEL","LINK",
    "PICTURE","ARTIST","TITLE","DURATION","RELEASE_DATE"
]

def _ensure_storage():
    if not DATA_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = DEFAULT_SHEET
        ws.append(HEADERS)
        wb.save(DATA_FILE)

def _write_records(records, sheet_name):
    _ensure_storage()

    wb = load_workbook(DATA_FILE)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]

    # kasujemy stare dane, zostawiamy nagłówki
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for rec in records:
        ws.append([rec.get(h, "") for h in HEADERS])

    wb.save(DATA_FILE)
    return DATA_FILE.stat().st_mtime

@app.route("/update_xlsx", methods=["POST"])
def update_xlsx():
    try:
        payload = request.get_json(force=True)
        records = payload.get("records", [])
        sheet = payload.get("sheetName", DEFAULT_SHEET)
        mtime = _write_records(records, sheet)
        return jsonify({
            "status": "ok",
            "message": "Zaktualizowano dane w Excelu",
            "file_name": DATA_FILE.name,
            "sheet_name": sheet,
            "updated_at": mtime
        })
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500

@app.route("/shutdown", methods=["POST"])
def shutdown():
    def _shutdown_later():
        time.sleep(0.5)
        os._exit(0)
    threading.Thread(target=_shutdown_later).start()
    return jsonify({"status": "shutting down"})

@app.route("/")
def index():
    return jsonify({"message": "server_4.py działa ✅"})

if __name__ == "__main__":
    print(f"Serwer uruchomiony: http://127.0.0.1:5000")
    print(f"Plik Excel: {DATA_FILE}")
    app.run(host="127.0.0.1", port=5000)
