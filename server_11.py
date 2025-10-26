from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import load_workbook, Workbook
from pathlib import Path
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)

BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "ZAJEBISTE_DANE.xlsx"
DEFAULT_SHEET = "Sheet1"
global_status_message = ""

# ======================================================
# Pomocnicze funkcje
# ======================================================

def _ensure_storage():
    """Upewnia się, że istnieje plik XLSX."""
    if not DATA_FILE.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = DEFAULT_SHEET
        ws.append(["SELECTOR","FOLDER","ADDED","LABEL","LINK",
                   "PICTURE","ARTIST","TITLE","DURATION","RELEASE_DATE"])
        wb.save(DATA_FILE)
        print("📁 Utworzono nowy plik ZAJEBISTE_DANE.xlsx")


def _update_excel(records, sheet_name):
    """Aktualizuje dane w arkuszu i tworzy kopie zapasowe."""
    global global_status_message
    _ensure_storage()

    wb = load_workbook(DATA_FILE, data_only=False)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    headers = [cell.value for cell in ws[1]]
    col_index = {h: i+1 for i, h in enumerate(headers) if h}

    # 🔁 Wpisywanie danych
    row = 2
    for rec in records:
        rec_upper = {k.upper(): v for k, v in rec.items()}
        for key, value in rec_upper.items():
            if key in col_index:
                ws.cell(row=row, column=col_index[key]).value = value
        row += 1

    # 🧹 Czyści puste wiersze (bez usuwania formatowania)
    for r in range(row, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).value = None

    try:
        # 📦 Zapisz główny plik
        wb.save(DATA_FILE)

        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        backup_name = f"{DATA_FILE.stem}_backup_{datetime.now():%Y-%m-%d_%H-%M-%S}.xlsx"
        backup_path = DATA_FILE.parent / backup_name
        wb.save(backup_path)

        # 🧹 Usuń stare backupy – zostaw maksymalnie 10 najnowszych
        backups = sorted(
            DATA_FILE.parent.glob(f"{DATA_FILE.stem}_backup_*.xlsx"),
            key=os.path.getmtime,
            reverse=True
        )
        for old_file in backups[10:]:
            try:
                os.remove(old_file)
                print(f"🧹 Usunięto stary backup: {old_file.name}")
            except Exception as cleanup_err:
                print(f"⚠️ Nie udało się usunąć {old_file.name}: {cleanup_err}")

        # 🧾 Wiadomość do konsoli i frontendu
        global_status_message = (
            f"✅ Plik '{DATA_FILE.name}' zaktualizowany {now_str}.\n"
            f"💾 Utworzono backup: {backup_path.name}"
        )

        print(global_status_message)

    except (PermissionError, OSError, IOError):
        # 📛 Excel otwarty lub plik zablokowany → zapis awaryjny
        backup_path = DATA_FILE.with_stem(
            DATA_FILE.stem + f"_temp_{datetime.now():%Y-%m-%d_%H-%M-%S}"
        )
        try:
            wb.save(backup_path)
            global_status_message = (
                f"💥 NIE UDAŁO SIĘ ZAPISAĆ DO '{DATA_FILE.name}'!\n"
                f"🔒 Plik jest otwarty w Excelu.\n"
                f"💾 Dane zapisano awaryjnie w '{backup_path.name}'.\n"
                f"❗ Zamknij Excela i kliknij AKTUALIZUJ ponownie."
            )
            print(global_status_message)
        except Exception as inner:
            global_status_message = (
                f"🚨 BŁĄD: Nie udało się zapisać pliku ani stworzyć kopii awaryjnej!\n"
                f"Szczegóły: {inner}"
            )
            print(global_status_message)

    return DATA_FILE.stat().st_mtime


# ======================================================
# Flask endpointy
# ======================================================

@app.route("/update_xlsx", methods=["POST"])
def update_xlsx():
    """Aktualizuje dane z frontendu."""
    try:
        payload = request.get_json(force=True)
        records = payload.get("records", [])
        sheet = payload.get("sheetName", DEFAULT_SHEET)
        mtime = _update_excel(records, sheet)

        return jsonify({
            "status": "ok",
            "message": global_status_message,
            "file_name": DATA_FILE.name,
            "updated_at": mtime
        })
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


@app.route("/get_xlsx", methods=["GET"])
def get_xlsx():
    """Zwraca dane z pliku XLSX do automatycznego wczytania na stronie."""
    try:
        wb = load_workbook(DATA_FILE, data_only=True)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                record = dict(zip(headers, row))
                records.append(record)

        # ✅ DODANE: zwracamy również nazwę pliku
        return jsonify({
            "status": "ok",
            "file_name": DATA_FILE.name,
            "sheet_name": ws.title,
            "records": records
        })

    except Exception as e:
        return jsonify({"status": "error", "error": str(e)}), 500


# ======= Endpoint testowy dla statusu serwera (dla strony HTML) =======
@app.route("/", methods=["GET"])
def index_status():
    from datetime import datetime
    return {
        "status": "ok",
        "server": "Flask",
        "message": "Serwer Flask działa poprawnie",
        "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }, 200


if __name__ == "__main__":
    print(f"🚀 server_10.py startuje na http://127.0.0.1:5000")
    try:
        app.run(host="127.0.0.1", port=5000)
    except KeyboardInterrupt:
        print("\n🧹 Serwer został zatrzymany przez użytkownika.")
