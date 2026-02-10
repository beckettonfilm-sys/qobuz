#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Qobuz multi-label scraper (2 pages max per label) with robust release-date handling.

INPUT
-----
- labels_scrapper.txt (same folder as this script)
  Each non-empty line format:
      Label Name - https://www.qobuz.com/us-en/label/.../download-streaming-albums/<id>?...

BEHAVIOR
--------
1) User provides:
   - date range OD/DO (DD.MM.RRRR), inclusive
   - minimum duration in minutes (albums shorter than this are rejected)
   - delays for listing pages and album pages

2) For each label from the file (max 2 listing pages per label):
   - scan listing page 1
   - if page 2 exists, scan listing page 2
   - for each album tile/link, extract listing release date using:
       - "Released by ... on Month D, YYYY"
       - "To be released on Month D, YYYY"
       - "To be released on M/D/YY" (rare on listings, common on album pages)
   - keep only albums whose LISTING release date is within the given range
     (strict: if a tile has no parseable listing date, it is skipped)

3) For each candidate album:
   - fetch album page and extract:
       album_title, main_artists, total length (Total length: HH:MM:SS)
   - ALSO extract album-page release date (same patterns as above)
   - if album-page release date is found, it becomes the source of truth and must be
     within the given range; otherwise we fall back to the listing date (already in range)
   - keep only albums where total length >= minimum minutes
   - keep only albums where the FIRST genre category in "About the album" is "Classical"

4) FINAL DEDUPLICATION (right before writing output files):
   - remove duplicates by (album_title, main_artists) within the same label
     (if label differs, both entries are kept)
   - duplicates are removed even if album_url is different

OUTPUT
------
- list_links.txt
  One album URL per line (after deduplication)

- title_artist_label.xlsx
  Columns (after deduplication):
    album_title | main_artists | label | album_url | release_date

Dependencies
------------
    pip install requests beautifulsoup4 rich openpyxl
"""

from __future__ import annotations

import random
import re
import sys
import time
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional, Tuple
from urllib.parse import urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.progress import (
    BarColumn,
    Progress,
    SpinnerColumn,
    TextColumn,
    TimeElapsedColumn,
    TimeRemainingColumn,
)
from rich.prompt import IntPrompt, Prompt


console = Console()

LABELS_FILE = "labels_scrapper.txt"
OUT_LINKS = "list_links.txt"
OUT_XLSX = "title_artist_label.xlsx"
OUT_MISSING_ALBUM_DATES = "album_date_missing.txt"

OUT_REJECTED_BY_GENRE = "rejected_by_genre.xlsx"
REQUEST_TIMEOUT = 20
RETRIES = 3

# Hard requirement: max 2 pages per label
MAX_PAGES_PER_LABEL = 2

# Album page contains:
# "Total length: 00:03:58"
RE_TOTAL_LENGTH = re.compile(r"Total length:\s*([0-9]{2}:[0-9]{2}:[0-9]{2})", re.IGNORECASE)

# Genre line (used for reference; extraction is from "About the album" section).
RE_GENRE = re.compile(r"\bGenre:\s*([^\n\r]+)", re.IGNORECASE)

# Release date patterns (listing and album pages)
RE_RELEASED_BY = re.compile(r"\bReleased by .*? on ([A-Za-z\.]+ \d{1,2}, \d{4})", re.IGNORECASE)
RE_RELEASED_ON = re.compile(r"\bReleased on ([A-Za-z\.]+ \d{1,2}, \d{4})", re.IGNORECASE)
RE_TO_BE_RELEASED_MONTH = re.compile(r"\bTo be released on ([A-Za-z\.]+ \d{1,2}, \d{4})", re.IGNORECASE)
RE_TO_BE_RELEASED_NUM = re.compile(r"\bTo be released on (\d{1,2}/\d{1,2}/\d{2,4})", re.IGNORECASE)
RE_RELEASED_BY_NUM = re.compile(r"\bReleased by .*? on (\d{1,2}/\d{1,2}/\d{2,4})", re.IGNORECASE)
RE_RELEASED_ON_NUM = re.compile(r"\bReleased on (\d{1,2}/\d{1,2}/\d{2,4})", re.IGNORECASE)


@dataclass(frozen=True)
class LabelSource:
    name: str
    url: str


@dataclass(frozen=True)
class Candidate:
    album_url: str
    label_name: str
    release_date_listing: date


@dataclass(frozen=True)
class AlbumDetails:
    title: str
    main_artists: str
    total_length_hms: str
    total_seconds: int
    release_date_album: Optional[date]
    genre_first: Optional[str]


@dataclass(frozen=True)
class OutputRecord:
    album_title: str
    main_artists: str
    label: str
    album_url: str
    release_date: date


def parse_pl_date(s: str) -> date:
    return datetime.strptime(s.strip(), "%d.%m.%Y").date()


def _norm_month_token(s: str) -> str:
    # remove trailing dot in abbreviations, e.g. "Jan." -> "Jan"
    return s.replace(".", "")


def parse_english_month_date(s: str) -> Optional[date]:
    s = " ".join((s or "").split()).strip()
    if not s:
        return None

    # normalize common oddities
    s = s.replace("Sept ", "Sep ")
    # remove dots in month abbreviations (Jan., Feb., etc.)
    parts = s.split(" ", 1)
    if parts:
        parts[0] = _norm_month_token(parts[0])
        s = " ".join(parts)

    for fmt in ("%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_numeric_us_date(s: str) -> Optional[date]:
    s = " ".join((s or "").split()).strip()
    if not s:
        return None

    for fmt in ("%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    # heuristic fallback: if it looks like dd/mm/yy (rare for us-en, but safe)
    for fmt in ("%d/%m/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    return None


def extract_release_date_from_text(text: str) -> Optional[date]:
    """Extract a release date from arbitrary text using Qobuz-like phrases.

    Supports both:
      - Month-name formats: "Jan 30, 2026"
      - Numeric formats (common on album pages, especially us-en): "1/30/26"
    """
    if not text:
        return None
    t = " ".join(text.split())

    # Month-name formats
    for rx in (RE_RELEASED_BY, RE_RELEASED_ON, RE_TO_BE_RELEASED_MONTH):
        m = rx.search(t)
        if m:
            return parse_english_month_date(m.group(1))

    # Numeric formats
    for rx in (RE_RELEASED_BY_NUM, RE_RELEASED_ON_NUM, RE_TO_BE_RELEASED_NUM):
        m = rx.search(t)
        if m:
            return parse_numeric_us_date(m.group(1))

    return None


def hms_to_seconds(hms: str) -> Optional[int]:
    m = re.fullmatch(r"(\d{2}):(\d{2}):(\d{2})", hms.strip())
    if not m:
        return None
    hh, mm, ss = map(int, m.groups())
    return hh * 3600 + mm * 60 + ss


def norm_key(s: str) -> str:
    """Normalization for dedup keys: strip, collapse whitespace, casefold."""
    return " ".join((s or "").split()).casefold()


def read_labels_file(path: Path) -> List[LabelSource]:
    if not path.exists():
        console.print(
            f"[bold red]Brak pliku:[/bold red] {path.name} (powinien być w tym samym folderze co skrypt)."
        )
        return []

    labels: List[LabelSource] = []
    for i, raw in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("#"):
            continue

        name, sep, url = line.partition(" - ")
        if not sep:
            console.print(f"[bold yellow]⚠️ Pomijam linię {i}:[/bold yellow] brak separatora ' - '")
            continue

        name = name.strip()
        url = url.strip()
        if not name or not url.startswith("http"):
            console.print(f"[bold yellow]⚠️ Pomijam linię {i}:[/bold yellow] niepoprawny format")
            continue

        labels.append(LabelSource(name=name, url=url))

    return labels


def normalize_label_base(url: str) -> str:
    """Remove trailing /page/<n> from label URL if present (keeps query)."""
    p = urlparse(url)
    path = re.sub(r"/page/\d+/?$", "", p.path)
    return urlunparse((p.scheme, p.netloc, path, p.params, p.query, p.fragment))


def build_label_page_url(label_url: str, page: int) -> str:
    p = urlparse(label_url)
    path = re.sub(r"/page/\d+/?$", "", p.path).rstrip("/")
    if page <= 1:
        new_path = path
    else:
        new_path = f"{path}/page/{page}"
    return urlunparse((p.scheme, p.netloc, new_path, p.params, p.query, p.fragment))


def make_session() -> requests.Session:
    s = requests.Session()
    s.headers.update(
        {
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X) QobuzLabelScraper/4.2",
            "Accept-Language": "en-US,en;q=0.9",
        }
    )
    return s


def polite_sleep(base_seconds: float, jitter_max: float = 0.25) -> None:
    if base_seconds <= 0:
        return
    time.sleep(base_seconds + random.uniform(0, max(0.0, jitter_max)))


def fetch_html(session: requests.Session, url: str) -> Optional[str]:
    last_err = None
    for attempt in range(1, RETRIES + 1):
        try:
            resp = session.get(url, timeout=REQUEST_TIMEOUT)

            if resp.status_code == 429:
                backoff = 8 + attempt * 6 + random.uniform(0, 4)
                console.print(f"[bold yellow]⏳ 429 Too Many Requests[/bold yellow] → czekam ~{backoff:.0f}s")
                time.sleep(backoff)
                continue

            if 500 <= resp.status_code < 600:
                backoff = 2 + attempt * 2 + random.uniform(0, 2)
                console.print(f"[bold yellow]⚠️ HTTP {resp.status_code}[/bold yellow] → retry za ~{backoff:.0f}s")
                time.sleep(backoff)
                continue

            if resp.status_code != 200:
                console.print(f"[bold yellow]⚠️ HTTP {resp.status_code}[/bold yellow] dla {url}")
                return None

            return resp.text

        except requests.exceptions.RequestException as e:
            last_err = e
            backoff = 1 + attempt * 2 + random.uniform(0, 2)
            console.print(
                f"[bold yellow]⚠️ Problem sieciowy[/bold yellow] (próba {attempt}/{RETRIES}) → retry za ~{backoff:.0f}s"
            )
            time.sleep(backoff)

    if last_err:
        console.print(f"[bold red]✖ Nie udało się pobrać[/bold red] {url} ({last_err})")
    return None


def listing_has_page2(soup: BeautifulSoup, label_url: str) -> bool:
    """Best-effort detection whether /page/2 exists in pagination."""
    base = normalize_label_base(label_url)
    base_path = urlparse(base).path.rstrip("/")
    pattern = re.compile(re.escape(base_path) + r"/page/2\b")

    for a in soup.find_all("a", href=True):
        href = a["href"]
        full = urljoin(base, href)
        if pattern.search(urlparse(full).path):
            return True

    # fallback: relaxed check
    for a in soup.find_all("a", href=True):
        if "/page/2" in a["href"]:
            return True

    return False


def extract_listing_release_date_for_link(a_tag, page_url: str, album_url: str, max_hops: int = 10) -> Optional[date]:
    """
    Walk up the DOM to find the smallest-ish container that:
      - contains a release-date phrase (Released by / To be released on / Released on)
      - contains ONLY this album link (not multiple different /album/ links)
    This prevents accidentally grabbing a date from a neighboring tile.
    """
    node = a_tag
    for _ in range(max_hops):
        if node is None:
            break

        # Collect distinct album URLs within this node
        try:
            album_hrefs = []
            for aa in node.find_all("a", href=True):
                href = aa["href"]
                if "/album/" in href:
                    album_hrefs.append(urljoin(page_url, href).split("#", 1)[0])
            distinct = set(album_hrefs)
        except Exception:
            distinct = set()

        # Must include our current album and not include other albums
        if distinct and album_url in distinct and len(distinct) == 1:
            try:
                txt = node.get_text(" ", strip=True)
            except Exception:
                txt = ""
            rel = extract_release_date_from_text(txt)
            if rel:
                return rel

        node = getattr(node, "parent", None)

    return None


def extract_album_candidates_from_listing(
    html: str,
    page_url: str,
    label_name: str,
    start: date,
    end: date,
) -> List[Candidate]:
    """Strict mode: if release date can't be read from listing, skip the album."""
    soup = BeautifulSoup(html, "html.parser")
    out: List[Candidate] = []
    seen_page_urls = set()

    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "/album/" not in href:
            continue

        album_url = urljoin(page_url, href).split("#", 1)[0]
        if album_url in seen_page_urls:
            continue
        seen_page_urls.add(album_url)

        rel = extract_listing_release_date_for_link(a, page_url=page_url, album_url=album_url)
        if not rel:
            continue

        if start <= rel <= end:
            out.append(Candidate(album_url=album_url, label_name=label_name, release_date_listing=rel))

    return out



def _clean_line_prefix(s: str) -> str:
    """Remove leading bullets/heading markers from a text line."""
    return re.sub(r"^[\s#*\-•]+", "", (s or "").strip()).strip()


def parse_album_first_genre(soup: BeautifulSoup) -> Optional[str]:
    """Extract the FIRST genre category from the *About the album* section.

    Qobuz often displays Genre twice:
      - near the top (often only a subgenre, e.g. "Chamber Music")
      - in the "About the album" block (usually the full taxonomy, e.g. "Classical / Chamber Music")

    This function reads ONLY the "About the album" block and returns its first
    category (e.g. "Classical").
    """
    raw_lines = [ln for ln in soup.get_text("\n", strip=True).split("\n") if ln.strip()]
    lines = [_clean_line_prefix(ln) for ln in raw_lines if _clean_line_prefix(ln)]

    # Find "About the album" heading (tolerate headings like "## About the album")
    about_idx: Optional[int] = None
    for i, ln in enumerate(lines):
        cf = ln.casefold()
        if cf == "about the album" or "about the album" in cf:
            about_idx = i
            break
    if about_idx is None:
        return None

    window = lines[about_idx : about_idx + 160]

    for j, ln in enumerate(window):
        cf = ln.casefold()
        if "genre" not in cf:
            continue

        raw = ""
        if ":" in ln:
            left, right = ln.split(":", 1)
            if left.strip().casefold() != "genre":
                continue
            raw = " ".join(right.split()).strip()
        else:
            if not cf.startswith("genre"):
                continue
            raw = " ".join(ln.split()[1:]).strip()

        if not raw:
            # Sometimes tags are on the next lines after a bare "Genre:"
            tags: List[str] = []
            for nxt in window[j + 1 : j + 10]:
                n = _clean_line_prefix(nxt)
                if not n:
                    continue
                # Stop on next field label
                if re.match(r"^(main artists|composer|label|total length|available in)\b", n, re.IGNORECASE):
                    break
                if n.endswith(":"):
                    break
                tags.append(n)
            if not tags:
                return None
            first = tags[0].strip()
            return first or None

        raw_cf = raw.casefold()
        if raw_cf.startswith("classical"):
            return "Classical"

        # Split by common delimiters used by Qobuz (in some locales)
        for delim in ("/", ",", "|", "›", ">"):
            if delim in raw:
                first = raw.split(delim, 1)[0].strip()
                return first or None

        # Fallback: first token
        first = raw.split(" ", 1)[0].strip()
        return first or None

    return None


def parse_album_release_date(soup: BeautifulSoup) -> Optional[date]:
    """Extract release date from album page.

    Qobuz often shows the release line near the top (bullets under the title), e.g.:
      - "Released on 1/30/26 by …"
      - "To be released on 2/27/26 by …"
      - "Released by … on Jan 30, 2026"

    We first scan individual lines (more precise), then fall back to whole-page text.
    """
    lines = [ln.strip() for ln in soup.get_text("\n", strip=True).split("\n") if ln.strip()]

    # Prefer early lines where the release info usually lives
    for ln in lines[:120]:
        if re.search(r"\b(released|to be released)\b", ln, re.IGNORECASE):
            d = extract_release_date_from_text(ln)
            if d:
                return d

    # Fallback: entire page text
    return extract_release_date_from_text(" ".join(lines))


def parse_album_details(html: str) -> Optional[AlbumDetails]:
    soup = BeautifulSoup(html, "html.parser")

    # Title: from H1, strip " by ..." if present
    title = ""
    h1 = soup.find("h1")
    if h1:
        t = h1.get_text(" ", strip=True)
        if " by " in t:
            title = t.split(" by ", 1)[0].strip()
        else:
            title = t.strip()

    # Main artists
    main_artists = ""
    main_block = None
    for tag in soup.find_all(["li", "p", "div"]):
        txt = tag.get_text(" ", strip=True)
        if txt.lower().startswith("main artists:"):
            main_block = tag
            break

    if main_block:
        artists = [aa.get_text(" ", strip=True) for aa in main_block.find_all("a") if aa.get_text(strip=True)]
        if artists:
            main_artists = ", ".join(artists).strip()
        else:
            main_artists = main_block.get_text(" ", strip=True).split(":", 1)[-1].strip()

    # Total length
    page_text = soup.get_text("\n", strip=True)
    m = RE_TOTAL_LENGTH.search(page_text)
    if not m:
        return None

    total_hms = m.group(1).strip()
    total_seconds = hms_to_seconds(total_hms)
    if total_seconds is None:
        return None

    release_date_album = parse_album_release_date(soup)
    genre_first = parse_album_first_genre(soup)

    if not title and not main_artists:
        return None

    return AlbumDetails(
        title=title,
        main_artists=main_artists,
        total_length_hms=total_hms,
        total_seconds=total_seconds,
        release_date_album=release_date_album,
        genre_first=genre_first,
    )


def write_links_txt(path: Path, links: List[str]) -> None:
    path.write_text("\n".join(links) + ("\n" if links else ""), encoding="utf-8")



def write_rejected_by_genre_xlsx(path: Path, rows: List[Tuple[str, str, str, str, str, str]]) -> None:
    """Write albums rejected by genre filter to an XLSX file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "rejected_by_genre"

    headers = ["album_title", "main_artists", "label", "album_url", "release_date", "genre_first"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    for (title, artists, label, url, rel, genre_first) in rows:
        ws.append([title, artists, label, url, rel, genre_first])

    widths = [40, 45, 28, 65, 14, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    wb.save(path)

def write_xlsx(path: Path, records: List[OutputRecord]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "albums"

    headers = ["album_title", "main_artists", "label", "album_url", "release_date"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)

    for r in records:
        ws.append(
            [
                r.album_title,
                r.main_artists,
                r.label,
                r.album_url,
                r.release_date.strftime("%d.%m.%Y"),
            ]
        )

    ws.freeze_panes = "A2"

    widths = [40, 45, 28, 65, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    wb.save(path)


def ask_float(prompt: str, default: float) -> float:
    while True:
        raw = Prompt.ask(prompt, default=str(default)).strip().replace(",", ".")
        try:
            v = float(raw)
            if v < 0:
                console.print("[bold red]Wartość nie może być ujemna.[/bold red]")
                continue
            return v
        except ValueError:
            console.print("[bold red]Podaj liczbę (np. 0.35).[/bold red]")


def main() -> None:
    console.print("[bold magenta]Qobuz multi-label scraper[/bold magenta]")
    console.print(
        "[dim]Filtr: data (listing + weryfikacja na album page) + minimalny czas (album page) + gatunek (pierwszy = Classical). Deduplikacja na końcu.[/dim]\n"
    )

    script_dir = Path(__file__).resolve().parent
    labels_path = script_dir / LABELS_FILE
    labels = read_labels_file(labels_path)
    if not labels:
        sys.exit(1)

    # User inputs
    while True:
        try:
            start_s = Prompt.ask("[bold]Zakres dat OD[/bold] (DD.MM.RRRR)").strip()
            end_s = Prompt.ask("[bold]Zakres dat DO[/bold] (DD.MM.RRRR)").strip()
            start_date = parse_pl_date(start_s)
            end_date = parse_pl_date(end_s)
            break
        except ValueError:
            console.print("[bold red]Nieprawidłowy format daty.[/bold red] Przykład: 24.01.2026\n")

    if start_date > end_date:
        console.print("[bold yellow]⚠️ OD jest później niż DO[/bold yellow] — zamieniam kolejność.")
        start_date, end_date = end_date, start_date

    min_minutes = IntPrompt.ask("[bold]Minimalna długość albumu[/bold] (minuty)", default=15)
    if min_minutes < 0:
        min_minutes = 0

    delay_list = ask_float("Opóźnienie między stronami listy (sek.)", default=0.35)
    delay_album = ask_float("Opóźnienie między stronami albumów (sek.)", default=0.55)

    console.print("\n[bold]Ustawienia:[/bold]")
    console.print(f"• Labels: [bold]{len(labels)}[/bold] (z {LABELS_FILE})")
    console.print(
        f"• Data: [bold]{start_date.strftime('%d.%m.%Y')}[/bold] → [bold]{end_date.strftime('%d.%m.%Y')}[/bold] (włącznie)"
    )
    console.print(f"• Minimalna długość: [bold]{min_minutes}[/bold] min (odrzuca krótsze)")
    console.print(f"• Max stron na label: [bold]{MAX_PAGES_PER_LABEL}[/bold]")
    console.print(f"• Delay list: [bold]{delay_list}[/bold]s, Delay album: [bold]{delay_album}[/bold]s\n")

    session = make_session()

    # Phase 1: scan listing pages (up to 2 per label)
    candidates: List[Candidate] = []
    seen_album_per_label = set()  # (album_url, label_name)

    total_steps_est = len(labels) * MAX_PAGES_PER_LABEL

    scan_progress = Progress(
        SpinnerColumn(),
        TextColumn("[bold cyan]Skanuję listingi label[/bold cyan]"),
        BarColumn(),
        TextColumn("{task.completed}/{task.total}"),
        TimeElapsedColumn(),
        console=console,
    )

    # Phase 2: collect accepted records
    accepted_records: List[OutputRecord] = []
    mismatch_date = 0
    missing_album_date = 0
    rejected_by_genre = 0
    missing_album_date_rows: List[str] = []
    rejected_by_genre_rows: List[Tuple[str, str, str, str, str, str]] = []

    try:
        with scan_progress:
            task_id = scan_progress.add_task("scan", total=total_steps_est)

            for src in labels:
                base_url = normalize_label_base(src.url)

                # Page 1
                page1_url = build_label_page_url(base_url, 1)
                html1 = fetch_html(session, page1_url)
                scan_progress.advance(task_id)

                has2 = False
                if html1:
                    found = extract_album_candidates_from_listing(html1, page1_url, src.name, start_date, end_date)
                    for c in found:
                        key = (c.album_url, c.label_name)
                        if key not in seen_album_per_label:
                            seen_album_per_label.add(key)
                            candidates.append(c)

                    soup1 = BeautifulSoup(html1, "html.parser")
                    has2 = listing_has_page2(soup1, base_url)

                polite_sleep(delay_list)

                # Page 2
                if has2 and MAX_PAGES_PER_LABEL >= 2:
                    page2_url = build_label_page_url(base_url, 2)
                    html2 = fetch_html(session, page2_url)
                    scan_progress.advance(task_id)

                    if html2:
                        found2 = extract_album_candidates_from_listing(html2, page2_url, src.name, start_date, end_date)
                        for c in found2:
                            key = (c.album_url, c.label_name)
                            if key not in seen_album_per_label:
                                seen_album_per_label.add(key)
                                candidates.append(c)

                    polite_sleep(delay_list)
                else:
                    # keep progress consistent with estimate
                    scan_progress.advance(task_id)

        console.print(f"[bold green]✓[/bold green] Kandydaci po dacie z listingu: [bold]{len(candidates)}[/bold]\n")

        # Phase 2: fetch details and filter by minimum length (AND re-check date from album page)
        details_progress = Progress(
            SpinnerColumn(),
            TextColumn("[bold cyan]Pobieram strony albumów[/bold cyan]"),
            BarColumn(),
            TextColumn("{task.completed}/{task.total}"),
            TimeElapsedColumn(),
            TimeRemainingColumn(),
            console=console,
        )

        with details_progress:
            task2 = details_progress.add_task("albums", total=len(candidates))

            for cand in candidates:
                html = fetch_html(session, cand.album_url)
                details_progress.advance(task2)

                if not html:
                    polite_sleep(delay_album)
                    continue

                det = parse_album_details(html)
                if not det:
                    polite_sleep(delay_album)
                    continue

                # Re-check release date from album page, if present
                rel_album = det.release_date_album
                if rel_album is None:
                    missing_album_date += 1
                    missing_album_date_rows.append(
                        f"{cand.label_name}\t{cand.album_url}\t{cand.release_date_listing.strftime('%d.%m.%Y')}\t{det.title}\t{det.main_artists}"
                    )
                    # Use listing date as fallback (strict per earlier agreement),
                    # but it's already within range.
                    rel_final = cand.release_date_listing
                else:
                    rel_final = rel_album
                    if not (start_date <= rel_album <= end_date):
                        mismatch_date += 1
                        polite_sleep(delay_album)
                        continue

                # Genre gate: accept ONLY if first genre category is "Classical"
                gf = (det.genre_first or "").strip()
                if not gf.casefold().startswith("classical"):
                    rejected_by_genre += 1
                    # collect for final report
                    rejected_by_genre_rows.append(
                        (
                            det.title,
                            det.main_artists,
                            cand.label_name,
                            cand.album_url,
                            rel_final.strftime('%d.%m.%Y'),
                            gf or "(missing)",
                        )
                    )
                    polite_sleep(delay_album)
                    continue

                if det.total_seconds >= min_minutes * 60:
                    accepted_records.append(
                        OutputRecord(
                            album_title=det.title,
                            main_artists=det.main_artists,
                            label=cand.label_name,
                            album_url=cand.album_url,
                            release_date=rel_final,
                        )
                    )

                polite_sleep(delay_album)

    except KeyboardInterrupt:
        console.print("\n[bold yellow]🟡 Przerwano Ctrl+C[/bold yellow] — zapisuję to, co już zebrane…")

    # Final deduplication: (title, artists) within same label
    before_dedup = len(accepted_records)
    seen = set()
    deduped: List[OutputRecord] = []
    for r in accepted_records:
        key = (norm_key(r.album_title), norm_key(r.main_artists), norm_key(r.label))
        if key in seen:
            continue
        seen.add(key)
        deduped.append(r)

    # Prepare outputs
    out_links_path = script_dir / OUT_LINKS
    out_xlsx_path = script_dir / OUT_XLSX

    links = [r.album_url for r in deduped]

    write_links_txt(out_links_path, links)
    write_xlsx(out_xlsx_path, deduped)

    # Optional debug: album pages where release date couldn't be parsed
    missing_path = script_dir / OUT_MISSING_ALBUM_DATES
    if missing_album_date_rows:
        header = "label\talbum_url\tlisting_release_date\talbum_title\tmain_artists\n"
        missing_path.write_text(header + "\n".join(missing_album_date_rows) + "\n", encoding="utf-8")

    console.print("[bold green]💾 Zapisano pliki:[/bold green]")
    console.print(f"• {OUT_LINKS}  ([dim]{len(links)} linków po deduplikacji[/dim])")
    console.print(f"• {OUT_XLSX}  ([dim]{len(deduped)} wierszy po deduplikacji[/dim])\n")

    console.print("[bold]Podsumowanie:[/bold]")
    console.print(f"• Labels w pliku: [bold]{len(labels)}[/bold]")
    console.print(f"• Kandydaci po dacie (listing): [bold]{len(candidates)}[/bold]")
    console.print(f"• Przeszło filtr długości ({min_minutes} min): [bold]{before_dedup}[/bold]")
    if before_dedup != len(deduped):
        console.print(
            f"• Usunięte duplikaty (ten sam tytuł+wykonawca w obrębie tej samej wytwórni): [bold]{before_dedup - len(deduped)}[/bold]"
        )
    if mismatch_date:
        console.print(
            f"• Odrzucone po weryfikacji daty na album page (poza zakresem): [bold]{mismatch_date}[/bold]"
        )
    if rejected_by_genre:
        console.print(
            f"• Odrzucone przez filtr gatunku (pierwszy != Classical): [bold]{rejected_by_genre}[/bold]"
        )
        if rejected_by_genre_rows:
            console.print(f"  [dim]Raport odrzuconych (gatunek) zapisano do: {OUT_REJECTED_BY_GENRE}[/dim]")
    if missing_album_date:
        console.print(
            f"• Albumy bez rozpoznanej daty na album page (użyto daty z listingu): [bold]{missing_album_date}[/bold]"
        )
        console.print(f"  [dim]Zapisano listę URL-i do: {OUT_MISSING_ALBUM_DATES}[/dim]")
    console.print("[dim]Gotowe.[/dim]")


if __name__ == "__main__":
    main()
